
from datetime import datetime
from typing import List
from tqdm import tqdm
import phonenumbers
import openpyxl
import psycopg2
import re


user_id = 59
company_id = 14
data_store = 23 # need to check these values

workbook = openpyxl.load_workbook(filename="Caypro_filter_details.xlsx", read_only=True, data_only=True)

sheet1 = workbook["OPPURTUNITY"]
sheet2 = workbook["CUSTOMERS"]
sheet3 = workbook["Feedback"]

def crete_jds(conn, jd_names: List[str]):
    jd_ids = []
    for jd_name in jd_names:
        with conn.cursor() as cursor:
            # create jd
            sql = f"""
                INSERT INTO t_jds (
                    uuid,    jd_name,    designation,    experience,
                    technical_skills,    non_technical_skills,    num_of_position,    notice_period,   
                    company_name,    company_url,    pref_location,    jd_status_value,
                    pseudo_jd_id,    last_profile_added_time,    text,    targets,
                    "offset",    total,    user_id,    assign_to_id,
                    cv_scraping_status,    is_scraping_complete_check,    cvs_last_fetch_time,    jd_type,
                    jd_embedding, created_at, updated_at, status
                )
                VALUES (
                    gen_random_uuid(), -- uuid
                    '{jd_name}', -- jd_name
                    '', -- designation
                    0.0, -- experience
                    '', -- technical_skills
                    '', -- non_technical_skills
                    0, -- num_of_position
                    '', -- notice_period
                    '', -- company_name
                    '', -- company_url
                    '', -- pref_location
                    'IN-PROGRESS', -- jd_status_value
                    '', -- pseudo_jd_id
                    '2024-06-12 12:00:00+00', -- last_profile_added_time
                    '', -- text
                    0, -- targets
                    0, -- offset
                    0, -- total
                    {user_id}, -- user_id (assuming you have a separate table for users)
                    '{user_id}', -- assign_to_id
                    'not_started', -- cv_scraping_status
                    'N', -- is_scraping_complete_check
                    NULL, -- cvs_last_fetch_time
                    'full-time', -- jd_type
                    '{{}}' , -- jd_embedding (empty JSON object)
                    '2024-06-12 12:00:00+00', --created_at
                    '2024-06-12 12:00:00+00', -- updated_at
                    1 -- status
                )RETURNING jd_id;
            """

            cursor.execute(sql)
            jd_id = cursor.fetchone()[0]
            jd_ids.append(jd_id)

            # crete record on jds assignee
            sql = f"""
                INSERT INTO t_jds_assign(jd_id, assignee_id, assignor_id, is_unassigned, created_at, updated_at, status)
                       VALUES ({jd_id}, {user_id}, {user_id}, 0, '2024-06-12 12:00:00+00', '2024-06-12 12:00:00+00', 1)
            """
            cursor.execute(sql)

    print(len(jd_ids),"jds are created")
    return jd_ids

def create_status(conn):
    status_dict = {}
    with conn.cursor() as cursor:
        cursor.execute(f"SELECT rank FROM t_candidate_journey WHERE company_id={company_id} ORDER BY rank::int DESC LIMIT 1")
        rank = int(cursor.fetchone()[0])

        sql = """
            INSERT INTO t_candidate_journey("value", "rank", company_id, user_id, created_type, created_at, updated_at, status)
            VALUES (%s, %s, %s, %s, 2, '2024-06-12 12:00:00+00', '2024-06-12 12:00:00+00', 1)
            RETURNING id
            """
        for s2_row in sheet2.iter_rows(min_row=2, values_only=True):
            status = s2_row[7]
            if status and status not in status_dict:
                rank += 1
                cursor.execute(sql, [status.strip(), rank, company_id, user_id])
                status_dict[status] = int(cursor.fetchone()[0])
            
        for s3_row in sheet3.iter_rows(min_row=2, values_only=True):
            status = s3_row[1]
            if status and status not in status_dict:
                rank += 1
                cursor.execute(sql, [status.strip(), rank, company_id, user_id])
                status_dict[status] = int(cursor.fetchone()[0])

    return status_dict

def get_related_rows_in_sheet2(opportunity_id):
    res = []
    flag = False
    for s2_row in sheet2.iter_rows(min_row=2, values_only=True):
        if s2_row[0] == opportunity_id:
            flag = True
            res.append(s2_row)
        elif(flag):
            break

    return res

def get_related_rows_in_sheet3(lead_id):
    res = []
    flag = False
    for s3_row in sheet3.iter_rows(min_row=2, values_only=True):
        if s3_row[0] == lead_id:
            flag = True
            res.append(s3_row)
        elif(flag):
            break

    return res

def save_email(cursor, cid, email, is_valid=True):
    sql = f"""
            SELECT cid FROM t_candidate_contact_details WHERE email ILIKE '{email}'
        """
    cursor.execute(sql)
    results = cursor.fetchone()

    if not results:
        try:
            insert_script = f"""
                INSERT INTO t_candidate_contact_details (
                        cid, email, is_details_valid,
                        created_at, updated_at, status
                    )
                VALUES(
                        {cid}, '{email}', {int(is_valid)}, 
                        '2024-03-07 12:00:00+00', '2024-03-07 12:00:00+00', 1 
                    )
            """
            cursor.execute(
                insert_script
            )

        except Exception as e:
            print(f"Error while inserting phone_number {email} into t_candidate_contact_details table: {e}")

def phone_number_formatter(phone: str):
    phone = re.sub(r'^0+', '', phone).strip() # removing leading zeros
    if len(phone) < 6 or len(phone) > 16:
        print("invalid" , phone)
        return phone, None, None, 0
    
    try:
        if "+" in phone:  
            parsed_number = phonenumbers.parse(phone)
            number = str(parsed_number.national_number)
            country_code = "+" + str(parsed_number.country_code)
            country_alpha_code = phonenumbers.region_code_for_number(
                parsed_number
            )
            return number, country_code, country_alpha_code, 1

        else:
            ph = "+" + phone
            parsed_number = phonenumbers.parse(ph, None)
            if phonenumbers.is_valid_number(parsed_number):
                number = str(parsed_number.national_number)
                country_code = "+" + str(parsed_number.country_code)
                country_alpha_code = phonenumbers.region_code_for_number(
                    parsed_number
                )
                return number, country_code, country_alpha_code, 1

        return phone, None, None, 0

    except phonenumbers.phonenumberutil.NumberParseException as exp:
        return phone, None, None, 0 
    
def save_phone(cursor, cid, phone):
    try:
        formatted_number = phone_number_formatter(phone)
        phone_number, country_code, alpha_code, is_valid = formatted_number
    
        insert_script = """
            INSERT INTO t_candidate_contact_details (
                    cid, phone_number, country_code, alpha_code,is_details_valid,   
                    created_at, updated_at, status
                )
            VALUES
                (
                %s, %s, %s, %s, %s, 
                '2024-06-11 12:00:00+00', '2024-06-11 12:00:00+00', 1 
                )
        """
        cursor.execute(insert_script, [cid, phone_number, country_code, alpha_code, is_valid])

    except Exception as e:
        print(f"Error while inserting phone_number {phone_number} into t_candidate_contact_details table: {e}")
        pass

def create_candidate(conn, cv_details):
    print(f"\ncompany name: {cv_details[1]} cv_name: {cv_details[2]} Lead_id: {cv_details[3]}	Designation: {cv_details[4]} Email: {cv_details[5]}	Phone: {cv_details[6]}")
    with conn.cursor() as cursor:
        sql = f"""
            INSERT INTO t_candidate_cvs (
                is_complete_profile, company_id, data_store_id,
                name, phone_number, email, image_url, raw_cv,  
                cv_format, segments, created_at, updated_at, status
            )
            VALUES (
                1, -- is_complete_profile
                {company_id}, -- company_id
                {data_store}, -- data_store_id
                '{cv_details[2]}', -- name
                '', -- phone_number
                '', -- email
                '', -- image_url
                '', -- raw_cv
                '', -- cv_format
                '{{}}'::jsonb, -- segments (empty JSON object)
                '2024-06-11 12:00:00+00',
                '2024-06-11 12:00:00+00', 
                1
            )RETURNING cid;
        """
        cursor.execute(sql)
        cid = cursor.fetchone()[0]

        # store email and phone in contact_details table
        email = cv_details[5]
        phone = cv_details[6]
        if email:
            save_email(cursor, cid, email.strip())
        if phone:
            save_phone(cursor, cid, phone)

        # store designation and company name in card table 
        company_name = cv_details[1]
        designation = cv_details[4]

        sql = f"""
            INSERT INTO t_candidate_cv_card (
                cid, current_des, current_company, previous_des, previous_company, college_name,
                skills,  degree, graduation_year, summary,  total_exp, preferred_location,
                location,  address, country_code, lat, lng, current_ctc, expected_ctc,  
                notice_periods, employment_type, is_open_to_work, source,  created_at, updated_at, status
            )
            VALUES (
                {cid}, -- cid_id (assuming you have a separate table for CandidateTable)
                '{designation or ""}', -- current_des
                '{company_name}', -- current_company
                '', -- previous_des
                '', -- previous_company
                '', -- college_name
                '', -- skills
                '', -- degree
                '', -- graduation_year
                '', -- summary
                0.0, -- total_exp
                '', -- preferred_location
                '', -- location
                '', -- address
                '', -- country_code
                0.0, -- lat
                0.0, -- lng
                '', -- current_ctc
                '', -- expected_ctc
                '', -- notice_periods
                '', -- employment_type
                0, -- is_open_to_work
                'EXCEL', -- source
                '2024-06-11 12:00:00+00',
                '2024-06-11 12:00:00+00', 
                1
            );
        """
        cursor.execute(sql)

        sql = f"""
            INSERT INTO t_candidate_media_account (
                cid, company_id, social_media_name, social_media_account,
                social_url, social_extended_url, created_at, updated_at, status
            )
            VALUES (
                {cid}, -- cid_id 
                {company_id}, -- company_id 
                'EXCEL', -- social_media_name
                '', -- social_media_account
                '', -- social_url
                '{{}}'::jsonb, -- social_extended_url
                '2024-06-11 12:00:00+00',
                '2024-06-11 12:00:00+00', 
                1
            );
        """
        cursor.execute(sql)

        sql = f"""
            INSERT INTO t_candidate_cvs_user_history (
                cid_id, user_id, created_at, updated_at, status
            )
            VALUES (
                {cid}, {user_id}, '2024-06-11 12:00:00+00', '2024-06-11 12:00:00+00', 1
            );
        """
        cursor.execute(sql)

        sql = f"""
            INSERT INTO t_candidate_sources (
                cid, source, created_at, updated_at, status
            )
            VALUES (
                1001, 'EXCEL', '2024-06-11 12:00:00+00', '2024-06-11 12:00:00+00', 1
            );
        """
        cursor.execute(sql)
        print(f"candidate related tables are created for cid-{cid}")

    return cid

def convert_date_format(date_string):
    # Parse the input date string
    input_format = "%b %d, %Y"
    input_date = datetime.strptime(date_string, input_format)

    # Convert to the desired output format
    output_format = "%Y-%m-%d %H:%M:%S.%f+00"
    output_date = input_date.strftime(output_format)
    return output_date

def create_mapping_and_tasks(conn, jd_id, cid, lead_id, status, status_dict, opp_name, date):
    with conn.cursor() as cursor:
        # store in cv_jd mappings
        sql=f"""
           INSERT INTO t_cv_jd_mappings(cid, jd_id, created_at, updated_at, status)
	              VALUES({cid}, {jd_id}, '2024-06-12 13:00:00+00', '2024-06-12 13:00:00+00', 1)	
        """
        cursor.execute(sql)

        # create status table record
        candidate_status = status_dict[status] if status in status_dict else 1 # change default status for manishi
        sql=f"""
           INSERT INTO t_candidate_status(jd_id, cid, candidate_status, created_at, updated_at, status)
                  VALUES ({jd_id}, {cid}, '{candidate_status}', '2024-06-12 13:00:00+00', '2024-06-12 13:00:00+00', 1)        
        """
        cursor.execute(sql)

        # create record in other tables

        # creating tasks tot tasks - 4573
        tasks = get_related_rows_in_sheet3(lead_id)
        for task in tasks:
            status_id = status_dict[task[1]]
            feedback = task[2]
            sql=f"""
                INSERT INTO t_scheduled_task (
                    jd_id, cid, user_id,
                    title, feedback, scheduled_date,
                    is_completed, description, cv_status,
                    created_at, updated_at, status
                )
                VALUES (
                    {jd_id}, {cid}, {user_id}, -- user_id 
                    '{opp_name}', -- title
                    '{feedback}', -- feedback
                    '{convert_date_format(date)}', -- scheduled_date
                    1, -- is_completed
                    '', -- description
                    {status_id}, -- cv_status_id
                    '2024-06-12 13:00:00+00',
                    '2024-06-12 13:00:00+00',
                    1 
                );
            """
            cursor.execute(sql)
        
    print("cv_jd mapping completed mapping")


if __name__=="__main__":
    conn = psycopg2.connect(
        host="127.0.0.1",
        port="5432",
        database="hire10xdb2",
        user="postgres",
        password="123456",
    )

    jd_names = ["Caypro Australia Automotive", "Caypro Australia health Care", "Caypro Middle East"]
    jd_ids = crete_jds(conn, jd_names)
    status_dict = create_status(conn)
    idx = 0

    for s1_row in sheet1.iter_rows(min_row=2, values_only=True):
        if not s1_row[0]:
            idx += 1
        if idx > 2:
            break
        
        jd_id = jd_ids[idx]
        opportunity_id = s1_row[0]
        for s2_row in get_related_rows_in_sheet2(opportunity_id):
            current_status = s2_row[7]
            lead_id = s2_row[3]
            cid = create_candidate(conn, s2_row)
            create_mapping_and_tasks(conn, jd_id, cid, lead_id, current_status, status_dict, s1_row[1], s2_row[9])

    conn.commit()
    conn.close()
    print("\nsuccessfully completed, without errors!!!!!!")
   

"""
SELECT * FROM t_jds WHERE jd_id > 807 ORDER BY jd_id DESC
SELECT * FROM t_jds_assign WHERE id > 968 ORDER BY id DESC
SELECT * FROM t_candidate_journey WHERE id > 244 AND company_id=1 ORDER BY id DESC
SELECT * FROM t_candidate_cvs WHERE cid > 37232 ORDER BY cid DESC
SELECT * FROM t_candidate_cv_card WHERE id > 35150 ORDER BY id DESC
SELECT * FROM t_candidate_media_account WHERE id > 38489 ORDER BY id DESC
SELECT * FROM t_candidate_contact_details WHERE id > 7498 ORDER BY id DESC
SELECT * FROM t_cv_jd_mappings WHERE id > 39026 ORDER BY id DESC
SELECT * FROM t_candidate_status WHERE id > 37566 ORDER BY id DESC
SELECT * FROM t_scheduled_task WHERE id > 24 ORDER BY id DESC
SELECT * FROM t_candidate_cvs_user_history WHERE id > 13810 ORDER BY id DESC
SELECT * FROM t_candidate_sources WHERE id > 35606 ORDER BY id DESC
------------------------------------------------------------------------------------------------
DELETE FROM t_cv_jd_mappings WHERE id > 39026
DELETE FROM t_candidate_cv_card WHERE id > 35150	
DELETE FROM t_candidate_contact_details WHERE id > 7498
DELETE FROM t_scheduled_task WHERE id > 24
DELETE FROM t_candidate_status WHERE id > 37566
DELETE FROM t_candidate_journey WHERE id > 244
DELETE FROM t_candidate_media_account WHERE id > 38489
DELETE FROM t_candidate_cvs_user_history WHERE id > 13810
DELETE FROM t_candidate_sources WHERE id > 35606
DELETE FROM t_candidate_cvs WHERE cid > 37232
DELETE FROM t_jds_assign WHERE id > 968
DELETE FROM t_jds WHERE jd_id > 807
"""   